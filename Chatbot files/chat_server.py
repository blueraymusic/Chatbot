from flask import Flask, request
import datetime
import json, requests, openai
#import numpy as np
import random, time,  webbrowser, os
from Levenshtein import distance
#from pygame import mixer  
from googlesearch import *
import openpyxl
#import keyboard
import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords
from nltk.probability import FreqDist
import re
import pandas as pd
import hashlib
from bs4 import BeautifulSoup
import urllib.parse
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity



nltk.download('stopwords')

"""
import importlib.machinery
sys.path.append("C://Users//minas//Desktop//vscode (Adel")
loader = importlib.machinery.SourceFileLoader("Chatbot_script", "C://Users//minas//Desktop//vscode (Adel\\Chatbot_script.py")
Chatbot_script = loader.load_module()
import Chatbot_script
"""


now = datetime.datetime.now()
heure = now.strftime('%H:%M %p')
print(heure)


h = """
<!DOCTYPE html>
<html>
<head>
  <meta charSet="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1, user-scalable=no" />
  <title>Chat</title>
   <style>      
   @import url('https://fonts.googleapis.com/css?family=Poppins:400,500,600,700&display=swap');

/* Change the background color of the bxody */
body {  
    background-color: #d9d9d9;
  }


  
  /* Style the main container element */
  .container {
    max-width: 800px;
    margin: 0 auto;
    padding: 20px;
  }
  .footer {
    position: absolute;
    bottom: 20;
    margin-top: 30px;
    width: 100%;
    /* Other styles for the footer go here */
  }
  
  /* Style the header elements */
  header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 20px;
  }
  
  header h1 {
    font-size: 36px;
    color: #333;
    margin: 0;
    text-align: center;
    font-family: 'Poppins', sans-serif;
  }
  
  header a {
    color: #333;
    text-decoration: none;
  }
  
  /* Style the chat area */
  .chat-area {
    display: flex;
    flex-direction: column;
    height: calc(100vh - 300px);
    overflow-y: auto;
    padding: 20px 0;
  }
  
  /* Style the message bubbles */
  .message-bubble {
    background-color: #fff;
    border-radius: 10px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    margin: 10px 0;
    max-width: 70%;
    padding: 15px 20px;
    position: relative;
  }
  
  /* Style the message text */
  .message-text {
    font-family: "Signifier", sans-serif;
    font-size: 16px;
    color: #333;
  }
  
  /* Style the message timestamps */
  .message-timestamp {
    font-family: "Sohne Mono", monospace;
    font-size: 14px;
    color: #999;
    position: absolute;
    right: 20px;
    bottom: 20px;
  }
  
  /* Style the input form and elements */
  .input-form {
    display: flex;
    align-items: center;
    justify-content: space-between;
    background-color: #fff;
    border-radius: 10px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    margin-top: 20px;
    padding: 10px 20px;
  }
  
  .input-form input {
    font-family: "Sohne Mono", monospace;
    font-size: 16px;
    color: #333;
    background-color: transparent;
    border: none;
    outline: none;
    flex-grow: 1;
  }
  
  .input-form button {
    font-family: "Sohne", sans-serif;
    font-size: 16px;
    color: #fff;
    background-color: #333;
    border: none;
    outline: none;
    border-radius: 5px;
    padding: 5px 10px;
    cursor: pointer;
  }
  
  .footer button {
    font-family: "Sohne", sans-serif;
    font-size: 16px;
    color: #fff;
    background-color: #333;
    border: none;
    outline: none;
    border-radius: 5px;
    padding: 5px 10px;
    cursor: pointer;

  }
  .footer button:hover {
    color: #fff;
    background-color: rgb(97, 95, 95);
  }

  
/* Style the chat area scrollbar */
.chat-area::-webkit-scrollbar {
    width: 10px;
    background-color: #f5f5f5;
  }
  
  .chat-area::-webkit-scrollbar-thumb {
    background-color: #333;
    border-radius: 10px;
  }
  
  /* Style the message bubbles for the user and AI */
  .user .message-bubble {
    align-self: flex-end;
    background-color: #333;
    color: #fff;
  }
  
  .ai .message-bubble {
    align-self: flex-start;
  }
  
  /* Style the input form placeholder text */
  .input-form input::placeholder {
    color: #999;
  }
  
  /* Style the input form button on hover */
  .input-form button:hover {
    background-color: #666;
  }
   
  .header {
            color: white;
          background-color: #000000;
            padding: 8px 0;
            margin-bottom: 5px;
            flex: none;
            position: relative;
            z-index: 12;
            background-size: cover;
            box-shadow: 0 1px 4px 0 rgba(0, 0, 0, 0.2);
            background-position: 50%; 
            border-radius: 10px
         }
         .intercom-reaction-picker {
            backface-visibility: hidden;
            padding: 12px;
            text-align: center;
            color: #777;
            background-color: #f0f3f5;
            margin: 60px -44px -44px -44px;
            border-radius: 8px; }
            @media (max-width: 1023px) {
              .intercom-reaction-picker {
                margin: 30px -16px -16px -16px; } }
            @media (max-width: 600px) {
              .intercom-reaction-picker {
                margin: 30px -8px -12px -8px;
                border-radius: 8px; } }
            .intercom-reaction-picker.intercom-reaction-picker-with-prompt {
              height: 74px; }
            .intercom-reaction-picker .intercom-reaction-prompt {
              padding-top: 9px;
              padding-bottom: 4px;
              text-align: center; }
            .intercom-reaction-picker .intercom-reaction {
              background: none;
              color: inherit;
              border: none;
              padding: 0;
              font: inherit;
              cursor: pointer;
              outline: none;
              width: 50px;
              height: 100%;
              display: inline-block;
              text-align: center;
              font-size: 32px;
              transition: transform 0.16s cubic-bezier(0.65, 0.61, 0.18, 1.8) 0.02s, filter 0.32s linear;
              cursor: pointer;
              transform-origin: 50% 60%; }
              .intercom-reaction-picker .intercom-reaction span {
                cursor: pointer;
                line-height: 55px; }
              .intercom-reaction-picker .intercom-reaction:hover, .intercom-reaction-picker .intercom-reaction:focus {
                transform: scale(1.32);
                transition: transform .04s; }
              .intercom-reaction-picker .intercom-reaction:active {
                transform: scale(1.4);
                transition: transform .04s; }
          
          .intercom-reaction-picker-reaction-selected .intercom-reaction {
            filter: grayscale(100%); }
          .intercom-reaction-picker-reaction-selected .intercom-reaction-selected {
            filter: grayscale(0%);
            transform: scale(1.32); }
            .intercom-reaction-picker-reaction-selected .intercom-reaction-selected:hover {
              transform: scale(1.32); }
            .intercom-reaction-picker-reaction-selected .intercom-reaction-selected:active {
              transform: scale(1.32); }
         
        </style>
  </head>
    <body class="header__lite">
      <header class="header">
    <div class="container header__container o__ltr" dir="ltr">
      <div class="content">
        <div class="mo o__centered o__reversed header__meta_wrapper">
          <div class="title">
            <style>
              @import url("https://fonts.intercomcdn.com/proxima-nova/proxima-nova-all.css");
              .title{
                  margin: 10 px;
                  font-family: "proxima-nova"; 
                  font-size: 20px;  
                  position: left;
              }
            </style>
            <p style="color:#ffffff">Chatbot - Main</p>
            </div>
          </div>
        </div>
        </div>
      </div>
    </div>
  </header>
</head>
<body>
  <div class="container">   
    <header>
      <h1> <a href="C://Users//minas//Desktop//vscode (Adel//Chatbot files//Intro page.html"> ChatBot 3 </a></h1>
    </header>
    <div class="chat-area">
      <div class="ai message-bubble">
        <p class="message-text">Hello! How can I help you today?</p>
        <p class="message-timestamp">10:15 AM</p>
      </div>
      <div class="user message-bubble">
        <p class="message-text">Hi there! Are you looking for information on a specific topic?</p>
        <p class="message-timestamp">10:16 AM</p>
      </div>
      <div class="ai message-bubble">
        <p class="message-text">I might sometimes generate incorrect information</p>
        <p class="message-timestamp">10:17 AM</p>
      </div>
    </div>
    <form class="input-form" action="http://127.0.0.1:5000/send-message" method="post">
      <input type="text" name="message" placeholder="Enter your message" />
      <button type="submit">Send</button>
    </form>
  </div>

<div class="intercom-reaction-picker" dir="ltr">
    
   <div class="intercom-reaction-prompt">Hope to see you again!</div>

     <button class="intercom-reaction" data-reaction-text="disappointed" tabIndex="0" aria-label="Disappointed Reaction">
       <span data-emoji="disappointed" title="Sad">
       </span>
     </button>
     <button class="intercom-reaction" data-reaction-text="neutral_face" tabIndex="0" aria-label="Neutral face Reaction">
       <span data-emoji="neutral_face" title="Neutral">
       </span>
     </button>
     <button class="intercom-reaction" data-reaction-text="smiley" tabIndex="0" aria-label="Smiley Reaction">
       <span data-emoji="smiley" title="Happy">
       </span>
     </button>
 </div>

</div>

</section>
</div>
   </div>
 </body>

//Using the prebuilt smileys
<script src="https://static.intercomassets.com/assets/help_center-4bd5f02b3f728fa6a162738af841aa465b18d1f29be03fbbffe96556a1787882.js" nonce="1GMML1nkQwfWnonk+RozP7Q2RE+B4NAnCF5yURxaRPk=">

</script>

</body>
</html>
""" 

with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//intents.json", "r") as f:
        intents = json.load(f)
        
app = Flask(__name__, template_folder='C:\\Users\\minas\\Desktop\\vscode (Adel\\Chatbot files')

'''
@app.route('/send-data', methods=['POST'])     
def send_data():
    if 'nom' in globals() and 'email' in globals():
      pass
    
    else:
      global nom 
      global email
      # Get the message from the request body
      nom = request.form['name']
      email = request.form['email']
      
    #print(f"Your name is {nom} and You email is {email}")
    return h
'''

@app.route('/sign-up', methods=['POST'])     
def sign_up():
    with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//Signup_er.html", 'r') as f:
      err = f.read()
     
    with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//loginstyle.css", 'r') as f:
      log_style = f.read()
      
    with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//log2.html", 'r') as f:
      log2 = f.read()
    
    with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//bootstrap_login.css", 'r') as f:
      bookstrap = f.read()

    
    
    log2 = log2.replace("{source}", bookstrap).replace("{logstyle}", log_style)
    err = err.replace("{source}", log_style)
    
    username = request.form['username']
    email = request.form['email']
    password = request.form['password']
    Re_password = request.form['Re_password']
    
    
    if password == Re_password and len(password) >= 8:
      password_hash = hashlib.sha256(password.encode()).hexdigest()
      password_hash_double = hashlib.sha256(password_hash.encode()).hexdigest()
      
      path = 'C://Users//minas//Desktop//vscode (Adel//Chatbot files//User_data.xlsx'
      
      header = ['Username', 'Email', 'Password']
      data = [username, email, password_hash_double]

      if os.path.exists(path): 
        pass
      else:
        # create a new workbook if file does not exist
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(header)
        workbook.save(path)
        
      df = pd.read_excel(path)
   
      match = df.loc[(df['Email'] == email)]
      
      if match.shape[0] > 0:
        return err
      else:

        try:
            # open the existing file
            workbook = openpyxl.load_workbook(path)
            sheet = workbook.active
        except FileNotFoundError:
            # create a new workbook if file does not exist
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(header)

        sheet.append(data)
        workbook.save(path)
        return log2 

      
    else:
      return err
      
@app.route('/logs', methods=['POST'])     
def log_in():
      
    with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//loginstyle.css", 'r') as f:
      log_style = f.read()
      
    with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//chatpage.html", 'r') as f:
      chatpage = f.read()
    
    with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//log2_err.html", 'r') as f:
      err = f.read()
    
    with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//bootstrap_login.css", 'r') as f:
      bookstrap = f.read()
      
    with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//chat.css", 'r') as f:
      chat_css = f.read()
    
    username = request.form['username']
    password = request.form['password']
    
    
    password_hash = hashlib.sha256(password.encode()).hexdigest()
    password = hashlib.sha256(password_hash.encode()).hexdigest()
    #import pandas as pd

    chatpage= chatpage.replace("{source}", chat_css)
    err = err.replace("{source}", bookstrap).replace("{logstyle}", log_style)
     
    df = pd.read_excel('C:\\Users\\minas\\Desktop\\vscode (Adel\\Chatbot files\\User_data.xlsx')
    match = df.loc[(df['Username'] == username) & (df['Password'] == password)]

    
    if match.shape[0] > 0:
      if 'nom' in globals() and 'email' in globals():
        pass
    
      else:
        global nom 
        global email 
        
        nom = request.form['username']
        df = pd.read_excel("C:\\Users\\minas\\Desktop\\vscode (Adel\\Chatbot files\\User_data.xlsx")
        row = df.loc[df['Username'] == nom]
        email = str(row['Email'].values[0])
        
        
        def decrypt(message):
            decrypted = ""
            for char in message:
                if char.isalpha():
                    shifted = ord(char) - 3
                    if char.isupper():
                        if shifted > ord('Z'):
                            shifted -= 26
                        elif shifted < ord('A'):
                            shifted += 26
                    else:
                        if shifted > ord('z'):
                            shifted -= 26
                        elif shifted < ord('a'):
                            shifted += 26
                    decrypted += chr(shifted)
                else:
                    decrypted += char
            return decrypted
          
        decrypted = decrypt('dixnfgvkumamlzph')
        
        #email send
        import smtplib
        from email.message import EmailMessage

        # create the message object
        msg = EmailMessage()
        msg['Subject'] = 'Welcome to Chatbot 🙂'
        msg['From'] = 'informativequick@gmail.com'
        msg['To'] = email
        msg.set_content(f'''
                        Dear {nom},

        I am your new chatbot, designed to assist you with any questions or concerns you may have. I am programmed to understand and respond to a variety of topics, from basic inquiries about our services to complex troubleshooting. I am here to make your experience with our company as smooth and seamless as possible.

        Please feel free to ask me anything, and I will do my best to provide you with a helpful response. Whether it's a question about our products or services, or simply a suggestion on how we can improve, I am here to listen.

        Thank you for choosing our company, and I look forward to assisting you in any way that I can.

        Best regards,

        Chatbot
                        ''')

        # send the email
        with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.ehlo()

            # login to your account
            smtp.login('informativequick@gmail.com', decrypted)

            # send the email
            smtp.send_message(msg)
        
        
    
      return chatpage
    else:
      return err
    
    
@app.route('/log1-in', methods=['POST'])     
def log1_in():
    
    with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//loginstyle.css", 'r') as f:
      log_style = f.read()
      
    with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//chatpage.html", 'r') as f:
      chatpage = f.read()
    
    with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//log2_err.html", 'r') as f:
      err = f.read()
    
    with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//bootstrap_login.css", 'r') as f:
      bookstrap = f.read()
      
    with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//chat.css", 'r') as f:
      chat_css = f.read()
    
    username = request.form['username']
    password = request.form['password']
    
    
    password_hash = hashlib.sha256(password.encode()).hexdigest()
    password = hashlib.sha256(password_hash.encode()).hexdigest()
    #import pandas as pd

    chatpage= chatpage.replace("{source}", chat_css)
    err = err.replace("{source}", bookstrap).replace("{logstyle}", log_style)
     
    df = pd.read_excel('C:\\Users\\minas\\Desktop\\vscode (Adel\\Chatbot files\\User_data.xlsx')
    match = df.loc[(df['Username'] == username) & (df['Password'] == password)]

    
    if match.shape[0] > 0:
      global nom 
      global email 
      
      nom = request.form['username']
      df = pd.read_excel("C:\\Users\\minas\\Desktop\\vscode (Adel\\Chatbot files\\User_data.xlsx")
      row = df.loc[df['Username'] == nom]
      email = str(row['Email'].values[0]) 
      return chatpage
    else:
      return err
      
      
    

@app.route('/code', methods=['POST'])     
def retrieve_code():
  
    source_code = """

<!doctype html>
<html lang="en" data-direction="ltr">
  <head>
    <meta charset="utf-8">
    <meta http-equiv="X-UA-Compatible" content="IE=edge">
    <title>Code Center</title>
    <meta name="description" content="">
    <meta name="viewport" content="width=device-width, initial-scale=1">
  <link rel="stylesheet" href="https://static.intercomassets.com/assets/help_center-ef66cd65cb3a4e4a29c89fb6080455cd7ce5e70980917a02cf959f8b71f94965.css" />
  <style>
    @import url('https://fonts.googleapis.com/css?family=Poppins:400,500,600,700&display=swap');
    .paper {
overflow: wrap;
word-wrap: break-word;
};
    *{
      margin: 0;
      padding: 0;
      box-sizing: border-box;
      font-family: 'Poppins', sans-serif;
    }

    #click{
      display: none;

    }
    label{
      position: fixed;
      right: 30px;
      bottom: 20px;
      height: 55px;
      width: 55px;
      background: -webkit-linear-gradient(left,  #a259ff, #377bf0);
      text-align: center;
      line-height: 55px;
      border-radius: 50px;
      font-size: 30px;
      color: #fff;
      cursor: pointer;
      z-index: 9999;

    }
    label i{
      position: fixed;
      top: 50%;
      left: 50%;
      transform: translate(-50%, -50%);
      transition: all 0.4s ease;
      z-index: 9999;
    }
    label i.fas{
      opacity: 0;
      pointer-events: none;
    }
    #click:checked ~ label i.fas{
      opacity: 1;
      pointer-events: auto;
      transform: translate(-50%, -50%) rotate(180deg);
    }
    #click:checked ~ label i.fab{
      opacity: 0;
      pointer-events: none;
      transform: translate(-50%, -50%) rotate(180deg);
    }
    .wrapper{
      position: fixed;
      right: 30px;
      bottom: 0px;
      max-width: 400px;
      z-index: 9999;
      background: #fff;
      border-radius: 15px;
      box-shadow: 0px 15px 20px rgba(0,0,0,0.1);
      opacity: 0;
      pointer-events: none;
      transition: all 0.6s cubic-bezier(0.68,-0.55,0.265,1.55);
    }
    #click:checked ~ .wrapper{
      opacity: 1;
      bottom: 85px;
      pointer-events: auto;
    }
    .wrapper .head-text{
      line-height: 60px;
      color: #fff;
      border-radius: 15px 15px 0 0;
      padding: 0 20px;
      font-weight: 500;
      font-size: 20px;
      background: -webkit-linear-gradient(left,  #a259ff, #377bf0);
    }
    .wrapper .chat-box{
      padding: 20px;
      width: 100%;
    }
    .chat-box .desc-text{
      color: #515365;
      text-align: center;
      line-height: 25px;
      font-size: 9px;
      font-weight: 500;
    }
    .chat-box form{
      padding: 10px 15px;
      margin: 20px 0;
      border-radius: 25px;
      border: 1px solid lightgrey;
    }
    .chat-box form .field{
      height: 50px;
      width: 100%;
      margin-top: 20px;
    }
    .chat-box form .field:last-child{
      margin-bottom: 15px;
    }
    form .field input,
    form .field button,
    form .textarea textarea{
      width: 100%;
      height: 100%;
      padding-left: 20px;
      border: 1px solid lightgrey;
      outline: none;
      border-radius: 25px;
      font-size: 16px;
      transition: all 0.3s ease;
    }
    form .field input:focus,
    form .textarea textarea:focus{
      border-color: #b1aeaf;
    }
    form .field input::placeholder,
    form .textarea textarea::placeholder{
      color: silver;
      transition: all 0.3s ease;
    }
    form .field input:focus::placeholder,
    form .textarea textarea:focus::placeholder{
      color: lightgrey;
    }
    .chat-box form .textarea{
      height: 70px;
      width: 100%;
    }
    .chat-box form .textarea textarea{
      height: 100%;
      border-radius: 50px;
      resize: none;
      padding: 15px 20px;
      font-size: 16px;
    }
    .chat-box form .field button{
      border: none;
      outline: none;
      cursor: pointer;
      color: #fff;
      font-size: 18px;
      font-weight: 500;
      background: -webkit-linear-gradient(left,  #a259ff, #377bf0);
      transition: all 0.3s ease;
    }
    .chat-box form .field button:active{
      transform: scale(0.97);
    }
</style>
<link rel="shortcut icon" href="C://Users//minas//Desktop//Bot_face.png" sizes= "42x42">

  </head>

  <body class="">
    <header class="header">
  <div class="container header__container o__ltr" dir="ltr">
    <div class="content">
      <div class="mo o__centered o__reversed header__meta_wrapper">
        <div class="mo__body header__site_name">
          <div class="header__logo">
            <a href="/en/">
                <!--Head logo but no logo yet-->
            </a>
          </div>
        </div>
        <div class="mo__aside">
          <div class="header__links">
<!-- head link but no link yet-->         
          </div>
        </div>
      </div>
          <h1 class="header__headline">Code Center</h1>
      <form action="http://127.0.0.1:5000/code" method="post" autocomplete="off" class="header__form search">
        <input type="text" autocomplete="off" class="search__input js__search-input o__ltr" placeholder="Search for Html & Css of any website..." tabindex="1" name="q" value="">
        <div class="search_icons">
          <button type="submit" class="search__submit o__ltr"></button>

      </form>
      </div>
    </div>
  </div>
</header>

    <div class="container">
      <div class="content educate_content"><section class="section">
    <div class="g__space">

      <a class="paper ">
        <div class="collection o__ltr">
          <div class="collection__photo">
            <svg role='img' viewBox='0 0 48 48'><g id="book-bookmark" stroke-width="2" fill="none" fill-rule="evenodd" stroke-linecap="round"><path d="M35 31l-6-6-6 6V7h12v24z"></path><path d="M35 9h6v38H11a4 4 0 0 1-4-4V5" stroke-linejoin="round"></path><path d="M39 9V1H11a4 4 0 0 0 0 8h12" stroke-linejoin="round"></path></g></svg>
          </div>
          <div class="collection_meta" dir="ltr">
            <h2 class="t__h3 c__primary">Getting Started</h2>
            <p class="paper__preview">Html & Css</p>
              
            <div class="avatar">
          
  <div class="avatar__photo avatars__images o__ltr">
  </div>
 
  <div class="avatar__info">
    <div>
      
    </div>
  </div>
</div>

          </div>
        </div>
        <br>
        <br>
        <div class="text"> 
        <xmp>{source}</xmp>
        <style>
        .wrap-text {
        width: 100px;
        word-wrap: break-word;
      }
       </style>
       </div>
      </a>
    </div>
     
</section>
</div>
    </div>

        <body >
      <input type="checkbox" id="click">
      <label for="click">
      <i class="fab fa-facebook-messenger"></i>
      <i class="fas fa-times"></i>
      </label>
          <div class="wrapper">
             <div class="head-text">
                Let's chat? - Chatbot
             </div>
             <div class="chat-box">
                <div class="desc-text">
                  <p id="error-message">Please fill out the form below to start chatting !</p>
                </div>
                <form class ="User_submit" action="http://127.0.0.1:5000/send-data" method="post" id="Form" >  <!--change later action="http://127.0.0.1:5000/send-data" method="post"-->
                   <div class="field">
                      <input type="email" placeholder="Your Email" name="name" id="email" required>
                   </div>
                   <div class="field">
                      <input type="password" placeholder="Your Password" name="password" id="password" required>
                   </div>
                   <div class="field textarea">
                      <textarea cols="30" rows="10" placeholder="Purpose (eg. learn, information, etc).." name ="purpose" required></textarea>
                   </div>
                   <div class="field">
                      <button type="submit">Start Chat</button>
                   </div>
                </form>
               
  
             </div>
          </div>
    </body>
           <!-- Form -->
           <script>
            const s = document.querySelector('#Form');
            const emailInput = document.getElementById("email");
            const passwordInput = document.getElementById("password");
      
            s.addEventListener('submit', (event) => {
                event.preventDefault();
                const email = emailInput.value;
                const password = passwordInput.value;
      
                if (password.length < 8) {
                  const errorMessageElement = document.getElementById('error-message');
                  errorMessageElement.textContent = "Invalid Password or Email, password must be at least 8 characters long !";
        // change this line later for security purposes 
                } else if (email =='admin@gmail.com' && password == 'admin1234'){
                  window.location.href = 'C://Users//minas//Desktop//vscode (Adel//Chatbot files//chatpage.html'
      
                }else {
                  window.location.href= 'C://Users//minas//Desktop//vscode (Adel//Chatbot files//Loginpage.html'
                }
             });
      
          </script>
    
 
    <footer class="footer">
  <div class="container">
    <div class="content">
      <div class="u__cf"  dir="ltr">
          <div class="footer__links">
              <ul class="footer__link-list footer__links__custom">
                  <li class="footer__link" data-footer-link-type="custom">
                    <a target="_blank" rel="nofollow noreferrer noopener" href="C://Users//minas//Desktop//vscode (Adel//Chatbot files//Intro page.html">Home</a>
                  </li>
                  <li class="footer__link" data-footer-link-type="custom">
                    <a target="_blank" rel="nofollow noreferrer noopener" href="#">Log In</a>
                  </li>
              </ul>
              <ul class="footer__link-list footer__links__social">
                  <li class="footer__link footer__link__twitter" data-footer-link-type="twitter">
                    <a target="_blank" rel="nofollow noreferrer noopener" href="https://twitter.com/">
                      <svg viewBox="0 0 32 32" version="1.0" xml:space="preserve" xmlns="http://www.w3.org/2000/svg" xmlns:xlink="http://www.w3.org/1999/xlink"><path d="M31.993,6.077C30.816,6.6,29.552,6.953,28.223,7.11c1.355-0.812,2.396-2.098,2.887-3.63  c-1.269,0.751-2.673,1.299-4.168,1.592C25.744,3.797,24.038,3,22.149,3c-3.625,0-6.562,2.938-6.562,6.563  c0,0.514,0.057,1.016,0.169,1.496C10.301,10.785,5.465,8.172,2.227,4.201c-0.564,0.97-0.888,2.097-0.888,3.3  c0,2.278,1.159,4.286,2.919,5.464c-1.075-0.035-2.087-0.329-2.972-0.821c-0.001,0.027-0.001,0.056-0.001,0.082  c0,3.181,2.262,5.834,5.265,6.437c-0.55,0.149-1.13,0.23-1.729,0.23c-0.424,0-0.834-0.041-1.234-0.117  c0.834,2.606,3.259,4.504,6.13,4.558c-2.245,1.76-5.075,2.811-8.15,2.811c-0.53,0-1.053-0.031-1.566-0.092  C2.905,27.913,6.355,29,10.062,29c12.072,0,18.675-10.001,18.675-18.675c0-0.284-0.008-0.568-0.02-0.85  C30,8.55,31.112,7.395,31.993,6.077z" /></svg>

</a>                  </li>
                  <li class="footer__link footer__link__linkedin" data-footer-link-type="linkedin">
                    <a target="_blank" rel="nofollow noreferrer noopener" href="https://linkedin.com/">
                      <svg viewBox="0 0 32 32" version="1.0" xml:space="preserve" xmlns="http://www.w3.org/2000/svg" xmlns:xlink="http://www.w3.org/1999/xlink"><rect height="23" width="7" y="9"/><path d="M24.003,9C20,9,18.89,10.312,18,12V9h-7v23h7V19c0-2,0-4,3.5-4s3.5,2,3.5,4v13h7V19C32,13,31,9,24.003,9z"/><circle cx="3.5" cy="3.5" r="3.5"/></svg>

</a>                  </li>
              </ul>
          </div>
      </div>
    </div>
  </div>
</footer>
</div>
<script>
      var elements = document.querySelectorAll('.text');
      for (var i = 0; i < elements.length; i++) {
        elements[i].style.wordWrap = "break-word";
        elements[i].style.overflow = wrap;
        elements[i].style.overflow = "visible";
         elements[i].style.maxWidth = "50px";
      }
    </script>

  </body>

</html>
"""
  
    global link
    
    link = request.form['q']

    # Get the source code
    try:
      from urllib.parse import urljoin
      
      r = requests.get(link).text
      
      '''
      fixed_lines = []
      lines = re.split(r'\n', r)
      for line in lines:
          if len(line) > 100:
              # Split the line into multiple lines
              split_lines = [line[i:i+90] for i in range(0, len(line), 90)]
              for split_line in split_lines:
                  fixed_lines.append(split_line)
          else:
              fixed_lines.append(line)
              
      r = "\n".join(fixed_lines)
      '''
      

      # send HTTP request to the website
      url = link
      r = requests.get(url).text

      response = requests.get(url)

      # parse HTML content using BeautifulSoup
      soup = BeautifulSoup(response.content, 'html.parser')

      # find all link tags with rel='stylesheet'
      css_links = soup.find_all('link', rel='stylesheet')

      # find all script tags with src containing 'bootstrap'
      bootstrap_links = soup.find_all('script', src=lambda src: src and 'bootstrap' in src)

      # download and print the contents of the CSS files
      for link in css_links:
          css_url = urljoin(url, link.get('href'))
          css_response = requests.get(css_url)
          css = css_response.content
          '''
          style = f"<style>{css}</style>\n"
          style += str(r)
          '''
          print("-----------------------CSS--\n")
          
          r = f"<style>{css}</style>\n{r}"

      # download and print the contents of the bootstrap files
      for link in bootstrap_links:
          bootstrap_url = urljoin(url, link.get('src'))
          bootstrap_response = requests.get(bootstrap_url)
          bookstrap = bootstrap_response.content
          """style = f"<style>{bookstrap}</style>\n"
          style += str(r)"""

          print("-----------------------Bookstrap--\n"
                )
          r = f"<style>{bookstrap}</style>\n{r}"

      
      source_code = source_code.replace("{source}", str(r))
    
    except:
      source_code = source_code.replace("{source}", "          [!} Impossible to retrive the Html & Css, link maybe incorrect. \n           Note: It should start by https/http.")

    return source_code
  


@app.route('/send-message', methods=['POST'])
def send_message():
    global message
    # Get the message from the request body
    message = request.form['message']
    
    #get answers
    Bot_Answers(message)
    message = answer
    # Generate a response
    response = generate_response(message)

    return response


def Bot_Answers(message):
    global answer
    # Open the intents.json file and read the contents
    #with open("C://Users//minas//Desktop//vscode (Adel//Chatbot files//intents.json", "r") as f:
     #   intents = json.load(f)

    # Extract the list of intents from the file
    intents_list = intents["intents"]

    # Create a mapping from tags to intents
    tag_to_intent = {}
    for intent in intents_list:
        tag_to_intent[str(intent["tag"])] = intent
    while True:
        check = False
        # Get user input
        user_input = str(message)
        user_input = user_input.lower()
        input_original = user_input
        
        
        try:
            #Generate Gpt answers 
            # Set the API key
            openai.api_key = "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"
            # Define the prompt
            prompt = user_input
            # Generate a response from GPT
            ai = openai.Completion.create(engine="text-davinci-002", prompt= prompt, max_tokens=1024)
            #ai = str(ai.choices[0].text)
            '''
            try:
              url = "https://api.writesonic.com/v2/business/content/content-rephrase?num_copies=1"
              payload = {
                  "tone_of_voice": "professional",
                  "content_to_rephrase": str(ai)
              }
              headers = {
                  "accept": "application/json",
                  "content-type": "application/json",
                  "X-API-KEY": "01544d40-19ab-4b72-a2d5-d0c78df8b5b7"
              }

              response = requests.post(url, json=payload, headers=headers)
              r = json.loads(response.text)
              ai = r[0]["text"]
            
            except:
              prompt= f"Paraphrase this: {ai}"
              ai = openai.Completion.create(engine="text-davinci-002", prompt= prompt, max_tokens=1024)
              
            '''
                    
        except:
            ai = "issue"
            print("[!] Impossible to generate GPT answers!")
            time.sleep(1)
        
        
        #Get all utterances
        utterances = []
        for intent in intents["intents"]:
            for pattern in intent["patterns"]:
                utterances.append(pattern)
        for i in utterances:
            Interpretation_input = min(utterances, key=lambda x: distance(user_input, x))
            
        #get the tag 
        for intent in intents["intents"]:
            if Interpretation_input in intent["patterns"]:
                tag = intent["tag"]
                break
        
        #setting the re-check
        re_check = False
  
        if 'issue' in ai.lower():
            re_answer = None
         
          
        if user_input.lower().startswith("prompt inject"):
            user_input = user_input.replace("prompt inject", '').replace(":", '').strip()
            
            def search_xlsx_file(filename, sheetname, search_value):
              wb = openpyxl.load_workbook(filename)
              sheet = wb[sheetname]
              tokenized_search_value = search_value.split()[0]
    
              for row in sheet.iter_rows():
                  In_row =  row[0].value.lower()
                  if In_row.startswith(tokenized_search_value):
                      return row[1].value
              
              for row in sheet.iter_rows():
                In_row =  row[0].value.lower()
                # Vectorize the texts
                vectorizer = CountVectorizer().fit_transform([In_row, search_value])
                # Compute cosine similarity
                rounding = round(cosine_similarity(vectorizer)[0][1], 1) 
                
                if rounding >= 0.7:
                  return row[1].value
                
              return None

            # j 
            filename = 'C:\\Users\\minas\\Desktop\\vscode (Adel\\Chatbot files\\gpt_injection_scripts.xlsx'
            sheetname = 'Sheet1'
            search_value = user_input.lower().replace("prompt inject", "").replace(":", "").strip()
            result = search_xlsx_file(filename, sheetname, search_value)

            if result is not None:
              re_answer = f"\n Inject the prompt:{result}"
            else:
              re_answer = f'None existent injection for : {search_value}'
            
            re_check = True 
            
        
        
            
        
        if 'google_search:' in user_input.lower():
            query = user_input.lower().replace("google_search:", "")
            
            #redirection message 
            re_answer = "\U0001F642 You've been redirected!"
            re_check = True 
          
            #chrome_path = r'C://Program Files (x86)//Microsoft//Edge//Application//msedge.exe'
            for url in search(query, tld="co.in", num=1, stop=1, pause=2):
              webbrowser.open("https://google.com/search?q=%s" % query)
        
        if 'redirect_to_webpage' in user_input.lower():
            query = user_input.lower().replace("redirect_to_webpage", "").replace(":", "").strip()
            
            #redirect message 
            re_answer = "\U0001F642 You've been redirected!"
            re_check = True 

           # Check if user input is a valid URL
            parsed_url = urllib.parse.urlparse(query)
            if parsed_url.scheme in ['http', 'https'] and parsed_url.netloc:
                webbrowser.open(query)
            else:
                # Search for the query and open the first result
                search_results = list(search(query, num=1, stop=1, pause=2))
                if search_results:
                    webbrowser.open(search_results[0])
                else:
                    # Try appending common domain extensions to the input and check if they exist
                    domain_extensions = ['.com', '.gov', '.ca', '.org', '.net', '.edu', '.mil', '.io', '.ai', '.app', '.co']
                    for ext in domain_extensions:
                        url = f"https://{query}{ext}"
                        try:
                            response = requests.head(url, timeout=5)
                            if response.status_code < 400:
                                webbrowser.open(url)
                                break
                        except:
                            pass
              
        # Check if user input matches any of the patterns
        response = None
        
        for tag, intent in tag_to_intent.items():
            if user_input in intent["patterns"]:
                
                #Get Bot answer first 
                response = random.choice(intent["responses"])
                check = True
     

        if check or re_check: 
          if ai != "issue":
            answer = re_answer
          
          elif response != None:
            answer = response
            
            
          else: 
             response = 'We are sorry to inform you that our website is currently experiencing a server problem. Our team is working hard to resolve the issue as soon as possible. Please check back later for updates. :)'
  
          
            
          return answer
            
        
        else:
            try:
                answer = str(ai.choices[0].text)
                def contains_code(code):
                  # Regular expression pattern for matching code keywords
                  pattern_ruby = r"(^|\s)(class|module|def|end|if|else|elsif|unless|case|when|while|until)(\s|$)"
      
                  pattern_html = "<[^<]+>" 
                  pattern_css = "[^{}]+\{[^}]+\}" 
                  
                  #this will return either true or false 
                  match_ruby = re.search(pattern_ruby, code)
        
                  match_html = re.search(pattern_html, code)
                  match_css= re.search(pattern_css, code)
                  
                  
                  #checking for each code (Ruby, Python, C++, Html and css)
                  global language
                  if match_ruby:
                      language = "Ruby"
                      return True
                    
                  elif match_html:
                      language = "Html"
                      return True
                  elif match_css:
                      language = "css"
                      return True
          
                  else: 
                      return False 
                  

              #catch user input
                valid_user = contains_code(user_input)
                valid_Output = contains_code(answer)
                
                code_presense = False 
                if valid_user == True :
                  answer = f'The text you provided is being interpreted as a code (Programing language) from {language} that can not yet be compiled by Chatbot. \nPlease let me know if you have any other question.'
                  code_presense = True         
                elif valid_Output == True:
                  answer = "I'm sorry, I'm not quite sure what you're trying to say with that input. Could you please provide more information or rephrase your question? I'll do my best to help you."
                  code_presense = True 
                
                               
                response = None
                
                if code_presense == False:
                  text = answer
                  sentences = sent_tokenize(text)
                  # Tokenize each sentence into words
                  words = [word_tokenize(sentence) for sentence in sentences]
                  # Remove stop words
                  stop_words = set(stopwords.words('english'))
                  words = [[word for word in sentence if word.lower() not in stop_words] for sentence in words]
                  # Perform word frequency analysis
                  all_words = [word for sentence in words for word in sentence]
                  fdist = FreqDist(all_words)
                  # Extract the most common words and phrases
                  common_words = fdist.most_common(10)
                  words = []
                  c_check = 0
                  for word, count in common_words:
                      if (bool(re.match('^[a-zA-Z]+$', word))) :
                        words.append(word)
                        c_check = 1
                  
                  if c_check==0:
                    word = ["not a word"]            

                  size = len(words)    
                  common_words = words[size - 1]
                  
                  # Open the intents.json file
                  with open('C://Users//minas//Desktop//vscode (Adel//Chatbot files//intents.json', 'r') as file:
                      data = json.load(file)

                  x = False
                  for intent in data['intents']:
                        if intent["tag"] in words:
                            intent["patterns"].append(user_input)
                            intent["responses"].append(answer)
                            x = True
                
                  if x == False:  
                      # Create a new intent object
                      new_intent = {
                          "tag": str(common_words),
                          "patterns": [],
                          "responses": []
                      }

                      # Append the new intent object to the intents list
                      data['intents'].append(new_intent)


                      # Save the changes to the intents.json file
                      with open('C://Users//minas//Desktop//vscode (Adel//Chatbot files//intents.json', 'w') as file:
                          json.dump(data, file, indent=4)



                  for intent in data['intents']:
                      if intent["tag"] == str(common_words):
                          user_input = user_input.lower().replace("google_search:", "").replace("redirect_to_webpage:", "")
                          intent["patterns"].append(user_input)
                          intent["responses"].append(answer)
                              
                  # Save the changes to the intents.json file
                  with open('C://Users//minas//Desktop//vscode (Adel//Chatbot files//intents.json', 'w') as file:
                      json.dump(data, file, indent=4)
                  # Check if the file exists
                  file_path = 'C:\\Users\\minas\\Desktop\\vscode (Adel\\Chatbot files\\Bot_data.xlsx'
                  if not os.path.exists(file_path):
                      # Create a new workbook
                      workbook = openpyxl.Workbook()
                      # Create a new worksheet
                      worksheet = workbook.active
                      # Add a header row to the worksheet
                      worksheet.append(['-- User ID','-- Email Address --','-- User Input --','-- Interpretation_input--','-- Bot Answers --', '-- AI Answers --'])
                      # Save the workbook
                      workbook.save(file_path)
                  else:
                      # Open the existing workbook
                      workbook = openpyxl.load_workbook(file_path)
                      # Get the worksheet
                      worksheet = workbook.active
                    
                  # chnage this, only keep it because YOure using admin to login
                  if 'nom' not in globals() and 'email' not in globals():
                    global nom
                    global email 
                    nom = 'admin'
                    email = 'admin@gmail.com'
              

                  try:
                      # Add some data to the worksheet
                      worksheet.append([nom, email, input_original, Interpretation_input,response, answer])
                      # Save the workbook
                      workbook.save(file_path)
                      break
                  except Exception as e:
                    try:
                      worksheet.append([nom, email, input_original, Interpretation_input,response, "Not avalaible"])
                      break
                    
                    except:
                      print(f"An error occurred: {e}")
                  break 
                
            except:
                pass
    
          
            
            
def generate_response(message):
    #html code
    html = '''
<!DOCTYPE html>
<html>
<head>
  <meta charSet="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1, user-scalable=no" />
  <title>Chat</title>
  <style>
  /* Change the background color of the body */
body {
    background-color: #d9d9d9;
  }
  
  /* Style the main container element */
  .container {
    max-width: 800px;
    margin: 0 auto;
    padding: 20px;
  }
  
  /* Style the header elements */
  header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 20px;
  }
  
  header h1 {
    font-family: "Sohne", sans-serif;
    font-size: 36px;
    color: #333;
    margin: 0;
    text-align: center;
  }
  
  header a {
    color: #333;
    text-decoration: none;
  }
  .footer button {
    font-family: "Sohne", sans-serif;
    font-size: 16px;
    color: #fff;
    background-color: #333;
    border: none;
    outline: none;
    border-radius: 5px;
    padding: 5px 10px;
    cursor: pointer;
    margin-top: 20px;

  }
  .footer button:hover {
    color: #fff;
    background-color: rgb(97, 95, 95);
  }
  
  /* Style the chat area */
  .chat-area {
    display: flex;
    flex-direction: column;
    height: calc(100vh - 300px);
    overflow-y: auto;
    padding: 20px 0;
  }
  
  /* Style the message bubbles */
  .message-bubble {
    background-color: #fff;
    border-radius: 10px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    margin: 10px 0;
    max-width: 70%;
    padding: 15px 20px;
    position: relative;
  }
  
  /* Style the message text */
  .message-text {
    font-family: "Signifier", sans-serif;
    font-size: 16px;
    color: #333;
  }
  
  /* Style the message timestamps */
  .message-timestamp {
    font-family: "Sohne Mono", monospace;
    font-size: 14px;
    color: #999;
    position: absolute;
    right: 20px;
    bottom: 20px;
  }
  
  /* Style the input form and elements */
  .input-form {
    display: flex;
    align-items: center;
    justify-content: space-between;
    background-color: #fff;
    border-radius: 10px;
    box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    margin-top: 20px;
    padding: 10px 20px;
  }
  
  .input-form input {
    font-family: "Sohne Mono", monospace;
    font-size: 16px;
    color: #333;
    background-color: transparent;
    border: none;
    outline: none;
    flex-grow: 1;
  }
  
  .input-form button {
    font-family: "Sohne", sans-serif;
    font-size: 16px;
    color: #fff;
    background-color: #333;
    border: none;
    outline: none;
    border-radius: 5px;
    padding: 5px 10px;
    cursor: pointer;
  }

  
/* Style the chat area scrollbar */
.chat-area::-webkit-scrollbar {
    width: 10px;
    background-color: #f5f5f5;
  }
  
  .chat-area::-webkit-scrollbar-thumb {
    background-color: #333;
    border-radius: 10px;
  }
  
  /* Style the message bubbles for the user and AI */
  .user .message-bubble {
    align-self: flex-end;
    background-color: #333;
    color: #fff;
  }
  
  .ai .message-bubble {
    align-self: flex-start;
  }
  
  /* Style the input form placeholder text */
  .input-form input::placeholder {
    color: #999;
  }
  
  /* Style the input form button on hover */
  .input-form button:hover {
    background-color: #666;
  }
       
  .header {
            color: white;
          background-color: #000000;
            padding: 15px 0;
            margin-bottom: 5px;
            flex: none;
            position: relative;
            z-index: 12;
            background-size: cover;
            box-shadow: 0 1px 4px 0 rgba(0, 0, 0, 0.2);
            background-position: 50%; 
            border-radius: 10px
         }
         .intercom-reaction-picker {
            backface-visibility: hidden;
            /* http://stackoverflow.com/a/15623914/411051 */
            padding: 12px;
            text-align: center;
            color: #777;
            background-color: #f0f3f5;
            margin: 60px -44px -44px -44px;
            border-radius: 8px; }
            @media (max-width: 1023px) {
              .intercom-reaction-picker {
                margin: 30px -16px -16px -16px; } }
            @media (max-width: 600px) {
              .intercom-reaction-picker {
                margin: 30px -8px -12px -8px;
                border-radius: 8px; } }
            .intercom-reaction-picker.intercom-reaction-picker-with-prompt {
              height: 74px; }
            .intercom-reaction-picker .intercom-reaction-prompt {
              padding-top: 9px;
              padding-bottom: 4px;
              text-align: center; }
            .intercom-reaction-picker .intercom-reaction {
              background: none;
              color: inherit;
              border: none;
              padding: 0;
              font: inherit;
              cursor: pointer;
              outline: none;
              width: 50px;
              height: 100%;
              display: inline-block;
              text-align: center;
              font-size: 32px;
              transition: transform 0.16s cubic-bezier(0.65, 0.61, 0.18, 1.8) 0.02s, filter 0.32s linear;
              cursor: pointer;
              transform-origin: 50% 60%; }
              .intercom-reaction-picker .intercom-reaction span {
                cursor: pointer;
                line-height: 55px; }
              .intercom-reaction-picker .intercom-reaction:hover, .intercom-reaction-picker .intercom-reaction:focus {
                transform: scale(1.32);
                transition: transform .04s; }
              .intercom-reaction-picker .intercom-reaction:active {
                transform: scale(1.4);
                transition: transform .04s; }
          
          .intercom-reaction-picker-reaction-selected .intercom-reaction {
            filter: grayscale(100%); }
          .intercom-reaction-picker-reaction-selected .intercom-reaction-selected {
            filter: grayscale(0%);
            transform: scale(1.32); }
            .intercom-reaction-picker-reaction-selected .intercom-reaction-selected:hover {
              transform: scale(1.32); }
            .intercom-reaction-picker-reaction-selected .intercom-reaction-selected:active {
              transform: scale(1.32); }
         
        </style>
</head>
<body>
<body class="header__lite">
      <header class="header">
    <div class="container header__container o__ltr" dir="ltr">
      <div class="content">
        <div class="mo o__centered o__reversed header__meta_wrapper">
          <div class="title">
            <style>
              @import url("https://fonts.intercomcdn.com/proxima-nova/proxima-nova-all.css");
              .title{
                  margin: 10 px;
                  font-family: "proxima-nova"; 
                  font-size: 20px;  
                  position: left;
              }
            </style>
            <p ><a style="color:#ffffff" href="C://Users//minas//Desktop//vscode (Adel//Chatbot files//Intro page.html">Chatbot - Reply</a></p>
            </div>
          </div>
        </div>
        </div>
      </div>
    </div>
  </header>
</head>
  <div class="container">   
    <header>
      <h1>Chat Bot 3</h1>
      
    </header>
    <div class="chat-area">
      <div class="ai message-bubble">
        <p class="message-text">{message}</p>
        <p class="message-timestamp">{heure}</p>
      </div>
    </div> 
    
    <form class="input-form" action="http://127.0.0.1:5000/send-message" method="post">
      <input type="text" name="message" placeholder="Enter your message" />
      <button type="submit">Send</button>
    </form> 
    
  </div>
</body>
</body>

<div class="intercom-reaction-picker" dir="ltr">
    
   <div class="intercom-reaction-prompt">Hope to see you again!</div>

     <button class="intercom-reaction" data-reaction-text="disappointed" tabIndex="0" aria-label="Disappointed Reaction">
       <span data-emoji="disappointed" title="Sad">
       </span>
     </button>
     <button class="intercom-reaction" data-reaction-text="neutral_face" tabIndex="0" aria-label="Neutral face Reaction">
       <span data-emoji="neutral_face" title="Neutral">
       </span>
     </button>
     <button class="intercom-reaction" data-reaction-text="smiley" tabIndex="0" aria-label="Smiley Reaction">
       <span data-emoji="smiley" title="Happy">
       </span>
     </button>
 </div>

</div>

</section>
</div>
   </div>
 </body>

//Using the prebuilt smileys
<script src="https://static.intercomassets.com/assets/help_center-4bd5f02b3f728fa6a162738af841aa465b18d1f29be03fbbffe96556a1787882.js" nonce="1GMML1nkQwfWnonk+RozP7Q2RE+B4NAnCF5yURxaRPk=">

</script>
</html>
'''
    
    # Your code here to generate a response based on the message
    #Output
    html = str(html)
    html = html.replace("{message}", str(message))
    html = html.replace("{heure}", heure)
    
    return html
  



if __name__ == '__main__':
    app.run(debug=True)
